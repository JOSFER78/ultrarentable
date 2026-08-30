"""Unit tests for Certified Master Catalog Export Service & Endpoints."""
import io
import openpyxl
from fastapi.testclient import TestClient
from services.api.app.main import app

client = TestClient(app)


def test_export_certified_csv_endpoint():
    response = client.get("/api/v2/certified/export/csv")
    assert response.status_code == 200
    assert "text/csv" in response.headers.get("content-type", "")
    assert "attachment; filename=\"catalogo_master_certificadas" in response.headers.get("content-disposition", "")
    
    content = response.text
    lines = content.strip().split("\n")
    assert len(lines) >= 1  # Header line
    header = lines[0]
    assert "Tipo" in header
    assert "ID" in header
    assert "Nombre" in header
    assert "Ruta" in header
    assert "Profit Factor (PF)" in header


def test_export_certified_xlsx_endpoint():
    response = client.get("/api/v2/certified/export/xlsx")
    assert response.status_code == 200
    assert "spreadsheetml.sheet" in response.headers.get("content-type", "")
    assert "attachment; filename=\"catalogo_master_certificadas" in response.headers.get("content-disposition", "")
    
    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert "Catálogo Master" in wb.sheetnames
    assert "Estrategias Certificadas" in wb.sheetnames
    assert "Meta-Estrategias" in wb.sheetnames
    
    ws = wb["Catálogo Master"]
    assert ws.max_row >= 1
    assert ws.cell(1, 1).value == "Tipo"
    assert ws.cell(1, 2).value == "ID"
