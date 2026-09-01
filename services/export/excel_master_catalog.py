"""Master Catalog Export Service for Certified Strategies and Meta-Strategies.

Generates structured CSV and Excel (.xlsx) exports adhering strictly to Zero-Mock / Real-Only evidence.
Includes:
- Item Type (ESTRATEGIA / META_ESTRATEGIA)
- ID & Name
- Route (ULTRA / FONDEO)
- Symbol / Components
- Timeframe
- Monthly Return %
- Sample Duration (months)
- OOS Months
- Trades Count
- Profit Factor (PF)
- Max Drawdown % (Max DD %)
- 11-Gate Audit Status
- Engine Version
- Hashes (Strategy/Portfolio SHA-256, Ledger SHA-256, Evidence Bundle SHA-256)
- Certification Timestamp (UTC)
"""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional
from sqlalchemy.orm import Session

from services.api.app.db.database import CandidateModel


def _extract_sample_duration_months(candidate: CandidateModel) -> Optional[float]:
    """Derive total sample duration in months from candidate scorecard duration_info."""
    from services.api.app.api.certified_summary_router import _scorecard

    sc = _scorecard(candidate)
    duration_info = sc.get("duration_info")
    if isinstance(duration_info, dict):
        total_m = duration_info.get("total_months") or duration_info.get("sample_duration_months")
        if isinstance(total_m, (int, float)) and total_m > 0:
            return float(total_m)
        is_m = duration_info.get("is_months")
        oos_m = duration_info.get("oos_months")
        if isinstance(is_m, (int, float)) and isinstance(oos_m, (int, float)):
            return float(is_m + oos_m)

    total_m = sc.get("sample_duration_months") or sc.get("total_months")
    if isinstance(total_m, (int, float)) and total_m > 0:
        return float(total_m)
    return None


def get_master_catalog_rows(db: Session, route: Optional[str] = None) -> List[Dict[str, Any]]:
    """Retrieve unified master catalog data rows for strategies and meta-strategies."""
    from services.api.app.api.certified_summary_router import (
        get_certified_strategies_endpoint,
        get_certified_meta_strategies_endpoint,
    )

    strategies = get_certified_strategies_endpoint(route=route, db=db, limit=5000)
    meta_strategies = get_certified_meta_strategies_endpoint(route=route, db=db)

    # Fetch candidates map for duration lookup
    candidate_rows = db.query(CandidateModel).all()
    candidate_map = {c.candidate_id: c for c in candidate_rows}

    rows: List[Dict[str, Any]] = []

    # Process certified strategies
    for st in strategies:
        strat_id = st.get("strategy_id", "")
        cand = candidate_map.get(strat_id)
        sample_months = _extract_sample_duration_months(cand) if cand else None

        rows.append({
            "type": "ESTRATEGIA",
            "id": strat_id,
            "name": st.get("name", ""),
            "route": st.get("route", ""),
            "symbol_or_components": st.get("symbol", ""),
            "timeframe": st.get("timeframe", ""),
            "monthly_return_pct": st.get("monthly_return"),
            "annual_return_pct": st.get("annual_return"),
            "win_rate_pct": st.get("win_rate_pct"),
            "sample_duration_months": sample_months,
            "oos_months": st.get("oos_months"),
            "trades_count": st.get("total_trades", 0),
            "profit_factor": st.get("profit_factor", st.get("oos_profit_factor")),
            "max_drawdown_pct": st.get("max_drawdown_pct"),
            "gate_audit_status": "11/11 APROBADO" if st.get("all_gates_pass") else "PENDIENTE",
            # W4.2: antes el fallback (para el caso improbable de que el dict no traiga la
            # clave) era el literal hardcodeado "5.4.0" -- con el motor vigente en 5.17.0
            # (services/engine_version.py) eso habría mentido en el catálogo de auditoría
            # afirmando un motor viejo concreto que no es el real. Sin dato real, se reporta
            # explícitamente que no hay dato, nunca un número inventado.
            "engine_version": st.get("engine_version") or "SIN_DATO",
            "strategy_hash": st.get("strategy_hash", ""),
            "ledger_hash": st.get("ledger_hash", ""),
            "evidence_bundle_hash": st.get("evidence_bundle_hash", ""),
            "certified_at_utc": st.get("certified_at_utc", ""),
        })

    # Process certified meta-strategies
    for ms in meta_strategies:
        components = ms.get("components", [])
        if isinstance(components, list) and components:
            comp_strs = [
                f"{c.get('symbol', c.get('strategy_id', ''))} ({round(float(c.get('weight', 0))*100, 1)}%)"
                for c in components if isinstance(c, dict)
            ]
            components_summary = ", ".join(comp_strs)
        else:
            components_summary = "Compuesto Multi-Activo"

        m_ret = ms.get("monthly_return")
        ann_ret = float(m_ret) * 12.0 if isinstance(m_ret, (int, float)) else None

        rows.append({
            "type": "META_ESTRATEGIA",
            "id": ms.get("meta_strategy_id", ms.get("portfolio_id", "")),
            "name": ms.get("name", ""),
            "route": ms.get("target_route", ms.get("route", "")),
            "symbol_or_components": components_summary,
            "timeframe": "MULTI",
            "monthly_return_pct": m_ret,
            "annual_return_pct": ann_ret,
            "win_rate_pct": None,
            "sample_duration_months": None,
            "oos_months": None,
            "trades_count": None,
            "profit_factor": ms.get("combined_profit_factor"),
            "max_drawdown_pct": ms.get("combined_max_drawdown_pct") or ms.get("max_drawdown_pct"),
            "gate_audit_status": "COMPONENTES_11/11_VERIFICADOS",
            # W4.2: mismo fix que arriba -- sin dato real, nunca un literal hardcodeado.
            "engine_version": ms.get("engine_version") or "SIN_DATO",
            "strategy_hash": ms.get("canonical_hash", ms.get("portfolio_hash", "")),
            "ledger_hash": ms.get("combined_ledger_hash", ms.get("canonical_hash", "")),
            "evidence_bundle_hash": ms.get("canonical_hash", ""),
            "certified_at_utc": ms.get("created_at", ""),
        })

    return rows


def build_master_catalog_csv(db: Session, route: Optional[str] = None) -> str:
    """Generate CSV string of the master catalog."""
    rows = get_master_catalog_rows(db, route=route)
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    writer.writerow([
        "Tipo",
        "ID",
        "Nombre",
        "Ruta",
        "Simbolo / Componentes",
        "Timeframe",
        "Retorno Mensual %",
        "Retorno Anual %",
        "Win Rate %",
        "Duracion Muestra (Meses)",
        "Meses OOS",
        "Total Trades",
        "Profit Factor (PF)",
        "Max Drawdown %",
        "Estado Auditoria 11 Gates",
        "Version Motor",
        "Hash Estrategia/Portafolio",
        "Hash Ledger",
        "Hash Evidencia",
        "Fecha Certificacion (UTC)",
    ])

    for r in rows:
        writer.writerow([
            r["type"],
            r["id"],
            r["name"],
            r["route"],
            r["symbol_or_components"],
            r["timeframe"],
            f"{r['monthly_return_pct']:.2f}" if isinstance(r['monthly_return_pct'], (int, float)) else "",
            f"{r['annual_return_pct']:.2f}" if isinstance(r['annual_return_pct'], (int, float)) else "",
            f"{r['win_rate_pct']:.2f}" if isinstance(r['win_rate_pct'], (int, float)) else "",
            f"{r['sample_duration_months']:.1f}" if isinstance(r['sample_duration_months'], (int, float)) else "",
            f"{r['oos_months']:.1f}" if isinstance(r['oos_months'], (int, float)) else "",
            r["trades_count"] if r["trades_count"] is not None else "",
            f"{r['profit_factor']:.2f}" if isinstance(r['profit_factor'], (int, float)) else "",
            f"{r['max_drawdown_pct']:.2f}" if isinstance(r['max_drawdown_pct'], (int, float)) else "",
            r["gate_audit_status"],
            r["engine_version"],
            r["strategy_hash"],
            r["ledger_hash"],
            r["evidence_bundle_hash"],
            r["certified_at_utc"],
        ])

    return output.getvalue()


def build_master_catalog_xlsx(db: Session, route: Optional[str] = None) -> bytes:
    """Generate structured Excel (.xlsx) file bytes of the master catalog with multi-tabs and formatting."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    
    # Setup styles
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="0F172A")
    sub_font = Font(name="Calibri", size=10, italic=True, color="475569")
    
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    headers = [
        "Tipo",
        "ID",
        "Nombre",
        "Ruta",
        "Símbolo / Componentes",
        "Timeframe",
        "Retorno Mensual %",
        "Retorno Anual %",
        "Win Rate %",
        "Duración Muestra (Meses)",
        "Meses OOS",
        "Total Trades",
        "Profit Factor (PF)",
        "Max Drawdown %",
        "Estado Auditoría 11 Gates",
        "Versión Motor",
        "Hash Estrategia/Portafolio",
        "Hash Ledger",
        "Hash Evidencia",
        "Fecha Certificación (UTC)",
    ]

    all_rows = get_master_catalog_rows(db, route=route)
    strat_rows = [r for r in all_rows if r["type"] == "ESTRATEGIA"]
    meta_rows = [r for r in all_rows if r["type"] == "META_ESTRATEGIA"]

    sheets_data = [
        ("Catálogo Master", all_rows),
        ("Estrategias Certificadas", strat_rows),
        ("Meta-Estrategias", meta_rows),
    ]

    # Remove default sheet
    wb.remove(wb.active)

    for sheet_title, data in sheets_data:
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True

        # Write Header
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[1].height = 28

        # Write Data
        for row_idx, r in enumerate(data, 2):
            ws.row_dimensions[row_idx].height = 20
            values = [
                r["type"],
                r["id"],
                r["name"],
                r["route"],
                r["symbol_or_components"],
                r["timeframe"],
                r["monthly_return_pct"],
                r["annual_return_pct"],
                r["win_rate_pct"],
                r["sample_duration_months"],
                r["oos_months"],
                r["trades_count"],
                r["profit_factor"],
                r["max_drawdown_pct"],
                r["gate_audit_status"],
                r["engine_version"],
                r["strategy_hash"],
                r["ledger_hash"],
                r["evidence_bundle_hash"],
                r["certified_at_utc"],
            ]

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = data_font
                cell.border = thin_border

                # Alignment & Formatting
                if col_idx in (1, 4, 6, 15, 16):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx in (7, 8, 9, 10, 11, 12, 13, 14):
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    if col_idx in (7, 8, 9, 14) and isinstance(val, (int, float)):
                        cell.number_format = '0.00"%"'
                    elif col_idx in (10, 11) and isinstance(val, (int, float)):
                        cell.number_format = "0.0"
                    elif col_idx == 12 and isinstance(val, (int, float)):
                        cell.number_format = "#,##0"
                    elif col_idx == 13 and isinstance(val, (int, float)):
                        cell.number_format = "0.00"
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column width
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(headers[col_idx - 1])
            for r_idx in range(2, len(data) + 2):
                val_str = str(ws.cell(row=r_idx, column=col_idx).value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
